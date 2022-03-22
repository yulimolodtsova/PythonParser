import fitz
import openpyxl
import camelot
import psycopg2
import easygui
import psycopg2
import pdfminer
import sys
from pdfminer.high_level import extract_text_to_fp
from bs4 import BeautifulSoup
if sys.version_info > (3, 0):
    from io import StringIO
else:
    from io import BytesIO as StringIO
from pdfminer.layout import LAParams

print ("Старт работы")
filename = easygui.fileopenbox(filetypes=["*.pdf"])
pdf_document = filename
doc = fitz.open(pdf_document)
page = doc.loadPage(0)
page_text = page.getText("text")
a = page_text.find("Направление")
b = page_text.find("организационно")
if a == -1:
    a = page_text.find("Квалификация:")
    b = page_text.find("Виды проф. деятельности:")
print(page_text)
print("a = ", a, " b = ", b)
data = page_text[a:b]
print ("\n", data)


book = openpyxl.Workbook()
sheet1 = book.create_sheet('list 1', 0)
sheet1.column_dimensions['A'].width = 28
sheet1.column_dimensions['C'].width = 28
sheet1.column_dimensions['B'].width = 105
sheet1['A1'] = "Направление"
i = data.find('0')
direction = data[i:i+8]
sheet1['B1'] = direction
m = b = 0
m = data.find('Магистерская')
if (m==-1):
    b = data.find('Профиль')
    j = b
    sheet1['A2'] = 'Профиль'
    print('Профиль')
else:
    sheet1['A2'] = "Магистерская программа"
    print('Мага')
    j = m
name_direction = data[i+9:j]
sheet1['C1'] = name_direction
a = data.find('Квалификация')
b = data.find('Форма')
qualification = data[a+13:b]
qualification = ''.join(qualification.split())
sheet1['A3'] = 'Квалификация'
sheet1['B3'] = qualification
a = data.find('Год начала')
form = data[b+16:a]
form = ''.join(form.split())
sheet1['A4'] = "Форма обучения"
sheet1['B4'] = form
sheet1['A5'] = "Год начала обучения"
b = data.find("Выпускающая")
date_of_start = data[a+21:b]
sheet1['B5'] = date_of_start
sheet1['A6'] = "Выпускающая кафедра"
a = data.find("Срок")
kafedra = data[b+21:a]
sheet1['B6'] = kafedra
sheet1['A7'] = "Срок обучения"
sheet1['A8'] = "Типы задач проф. деятельности"
if (form == "Очная"):
    term = data[a+15:a+21]
    a = a+21
else:
    if(qualification =="бакалавр"):
        term = data[a+15:a+21]
        a = a + 21
    else:
        term = data[a+15:a+32]
        a = a+31
sheet1['B7'] = term
b = data.find("Виды")
if (b==-1):
    b = data.find("Типы")
prog = data[a:b]
if(prog[0]=='в'):
    prog = prog[1:]
sheet1['B2'] = prog

if (qualification == "Магистр"):
    tables = camelot.read_pdf(filename, pages = '2,3,4')
    kol = 3
elif (qualification == "бакалавр"):
    tables = camelot.read_pdf(filename, pages = '2,3,4,5,6')
    kol = 5
elif (qualification == "Инженер"):
    tables = camelot.read_pdf(filename, pages = '2,3,4,5,6,7,8')
    kol = 7
print (tables[0].parsing_report)
tables.export("tables.xlsx", f='excel')


wb = openpyxl.load_workbook("tables.xlsx")
ws = wb["page-2-table-1"]

mr = 5
mc = 25

for i in range (2, mr+1):
	for j in range (1, mc+1):
	 c = ws.cell(row = i, column = j+1).value
	 #print("c = ", c, "\n")
	 sheet1.cell(row = i+11, column = j).value = c
	 #print(" cel = ", sheet1.cell(row = i+12, column = j).value, "\n")

mr = 15
start = 8
i2 = 15
for t in range (0, kol):
    i = 7
    if (t == 0):
        ws = wb["page-2-table-1"]
    if (t == 1):
        ws = wb["page-3-table-1"]
    if (t == 2):
        ws = wb["page-4-table-1"]
    if (t == 3):
        ws = wb["page-5-table-1"]
    if (t == 4):
        ws = wb["page-6-table-1"]

    while ((ws.cell(row = i, column = 4).value!=None)or(ws.cell(row = i, column = 3).value!=None)):
        for j in range (1, mc+1):
             c = ws.cell(row = i, column = j+1).value
             if (c!=None):
                 c2 = str(c)
                 c = c2.replace('\n',' ')
                 sheet1.cell(row = i2, column = j).value = c
        i += 1
        i2 += 1
        #start = i + start + 1

book.save("resoult.xlsx")

#ПОДКЛЮЧЕНИЕ К БД
conn = psycopg2.connect(dbname="LK", user="postgres",
                        password="123", host="localhost", port=5432)
cur = conn.cursor()
#ПОДКЛЮЧЕНИЕ К БД

wb = openpyxl.load_workbook("resoult.xlsx")
ws = wb["list 1"]

#НАПРАВЛЕНИЕ
direction_code = ws.cell(row = 1, column = 2).value
print(direction_code)
cur.execute("SELECT direction_id FROM direction WHERE code = %s",(direction_code, ))
a = cur.fetchall()
if not a:
    print("Список пуст") #заносим направление в бд
    direction_name = ws.cell(row = 1, column = 3).value
    cur.execute("SELECT COUNT(*) FROM direction")
    kol = cur.fetchall()
    direction_id = kol[0][0] + 1
    print(kol[0][0])
    cur.execute("INSERT INTO direction (code, name, direction_id) VALUES (%s, %s, %s)",
    (direction_code.strip(), direction_name.strip(), direction_id))
    conn.commit()
else:
    print("Удаляем запись")#заносим направление в бд
    direction_id = a[0][0]
#НАПРАВЛЕНИЕ 
    
#ПРОФИЛЬ
    
profile_name = str(ws.cell(row = 2, column = 2).value)
print(profile_name)
cur.execute("SELECT profile_id FROM profile WHERE name = %s",(profile_name.strip(), ))
a = cur.fetchall()
if a :
    print("Такой профиль существует")
    profile_id = a[0][0]
else:
    cur.execute("SELECT COUNT(*) FROM profile")
    kol = cur.fetchall()
    profile_id = kol[0][0] + 1
    cur.execute("INSERT INTO profile (name, profile_id) VALUES (%s, %s)",
    (profile_name.strip(), profile_id))
    conn.commit()
    

#ПРОФИЛЬ
    
#УЧ. ПЛАН
st = ws.cell(row = 6, column = 2).value
year = ws.cell(row = 5, column = 2).value
training_form = ws.cell(row = 4, column = 2).value
qualification = ws.cell(row = 3, column = 2).value
st = ''.join(st.split())
training_form = ''.join(training_form.split())
qualification = ''.join(qualification.split())
kafedra_name = st[0:2]
print("Кафедра - ", kafedra_name)
if training_form == "Очная":
    form_of_training_id = 1
else:
    form_of_training_id = 2
    
if qualification == "Магистр":
    qualification_id = 2
elif qualification == "бакалавр" :
    qualification_id = 1
elif qualification == "Инженер" :
    qualification_id = 3
cur.execute("SELECT department_id FROM department WHERE short_name = %s",(kafedra_name, ))
a = cur.fetchall()
department_id = a[0][0]
cur.execute("SELECT syllabus_id FROM syllabus WHERE year = %s AND department_id = %s AND direction_id = %s AND profile_id = %s AND form_of_training_id = %s AND qualification_id = %s ",
            (year, department_id, direction_id, profile_id, form_of_training_id, qualification_id ))
a = cur.fetchall()
if a:
    syllabus_id = a[0][0]
else:
    cur.execute("SELECT COUNT(*) FROM syllabus")
    kol = cur.fetchall()
    syllabus_id = kol[0][0] + 1
    cur.execute("INSERT INTO syllabus (syllabus_id, year, department_id, direction_id, profile_id, qualification_id, form_of_training_id ) VALUES (%s, %s, %s, %s, %s, %s, %s)",
    (syllabus_id, year,department_id, direction_id, profile_id, qualification_id, form_of_training_id ))
    conn.commit()
#УЧ. ПЛАН

#ПОИСК НЕНУЖНЫХ ПРЕДМЕТОВ
pages = [2, 3, 4, 5, 6]
outf = open("1.html", "w")
with open(pdf_document, 'rb') as fin:
    extract_text_to_fp(fin, outf, laparams=LAParams(), output_type='html', codec=None, page_numbers = pages)
soup = BeautifulSoup(open("1.html"))
bTags = []
for i in soup.find_all('span', style=lambda x: x and 'Italic' in x):
    bTags.append(i.text)
str_x = ""
x = []
for i in bTags:
    #print (i)
    if (i.find("ПРАКТИКА")==-1):
        str_x = i.replace('\n','')
        str_x = str_x.replace(' ','')
        x.append(str_x)

print (x)

#ПОИСК НЕНУЖНЫХ ПРЕДМЕТОВ
    
#ДИСЦИПЛИНА И СОДЕРЖАНИЕ УЧЕБНОГО ПЛАНА
ii = 15
dop_name = ""
enter_in_db = True
at_f = []
while ((ws.cell(row = ii, column = 2).value!=None) or (ws.cell(row = ii, column = 3).value!=None)):
    if((ws.cell(row = ii, column = 2).value==None) or (ws.cell(row = ii, column = 3).value==None)):
        ii+=1
        continue
    discipline_name = ws.cell(row = ii, column = 2).value
    dop_name = discipline_name.replace(' ','')
    for i in x:
        if(i == dop_name):
            enter_in_db = False
        
    department_name = ws.cell(row = ii, column = 3).value
    department_name = department_name.replace('-', "")
    department_name = department_name.replace(' ', "")
    department_name = department_name.replace('\n','')
    department_name = department_name.replace("None","")
    
    if (enter_in_db):
        print(discipline_name, " - ", department_name)
        cur.execute("SELECT department_id FROM department WHERE short_name = %s",(department_name,  ))
        a = cur.fetchall()
        print(department_name)
        department_id = a[0][0]
        cur.execute("SELECT discipline_id FROM discipline WHERE name = %s AND department_id = %s",(discipline_name, department_id,))
        a = cur.fetchall()
        if a:
            discipline_id = a[0][0]
        else:
            cur.execute("SELECT COUNT(*) FROM discipline")
            kol = cur.fetchall()
            discipline_id = kol[0][0] + 1
            cur.execute("INSERT INTO discipline (discipline_id, name, department_id) VALUES (%s, %s, %s)",(discipline_id, discipline_name, department_id ))
            conn.commit()
        
        at_f.clear()
        for j in range (4 ,9):
            st = str(ws.cell(row = ii, column = j).value)
            st = st.replace('-', "")
            st = st.replace(' ', "")
            st = st.replace('\n','')
            st = st.replace("None","")
            st = st.split(',')
            at_f.append(st)
            
        print("Список: ", at_f)


        if at_f[0][0] =='' and at_f[1][0]=='' and at_f[2][0]=='' and at_f[3][0]=='' and at_f[4][0]=='':
            attestation_form_id = 8
            semester_number = 0
            cur.execute("SELECT syllabus_content_id FROM syllabus_content WHERE semester_number = %s AND attestation_form_id = %s AND discipline_id = %s AND syllabus_id = %s",
                            (semester_number, attestation_form_id, discipline_id, syllabus_id,))
            a = cur.fetchall()
            if not a:
                cur.execute("SELECT COUNT(*) FROM syllabus_content")
                kol = cur.fetchall()
                syllabus_content_id = kol[0][0] + 1
                cur.execute("INSERT INTO syllabus_content (syllabus_content_id, semester_number, attestation_form_id, discipline_id, syllabus_id  ) VALUES (%s, %s, %s, %s, %s)",
                (syllabus_content_id, semester_number, attestation_form_id, discipline_id, syllabus_id ))
                conn.commit()
            ii+=1
            continue

        if at_f[3][0]!='':
            g = set(tuple(at_f[3]))
            print(a)
            for i in range (0, 3):
                b = set(tuple(at_f[i]))
                c = g & b
                if c:
                    kol = len(c)
                    m = list(c)
                    for j in range (0, kol):
                        #st = at_f[4].replace(m[j], "")
                        #at_f[4] = st
                        #st = at_f[i].replace(m[j], "")
                        #at_f[i] = st
                        at_f[3].remove(m[j])
                        at_f[i].remove(m[j])
                        semester_number = m[j]
                        if i == 0:
                            attestation_form_id = 5
                        if i == 1:
                            attestation_form_id = 9
                        if i == 2:
                            attestation_form_id = 10
                        
                        cur.execute("SELECT syllabus_content_id FROM syllabus_content WHERE semester_number = %s AND attestation_form_id = %s AND discipline_id = %s AND syllabus_id = %s",
                            (semester_number, attestation_form_id, discipline_id, syllabus_id,))
                        a = cur.fetchall()
                        if not a:
                            cur.execute("SELECT COUNT(*) FROM syllabus_content")
                            kolvo = cur.fetchall()
                            syllabus_content_id = kolvo[0][0] + 1
                            cur.execute("INSERT INTO syllabus_content (syllabus_content_id, semester_number, attestation_form_id, discipline_id, syllabus_id  ) VALUES (%s, %s, %s, %s, %s)",
                                (syllabus_content_id, semester_number, attestation_form_id, discipline_id, syllabus_id ))
                            conn.commit()
               
            
        if at_f[4][0]!='':
            g = set(tuple(at_f[4]))
            print(a)
            for i in range (0, 3):
                b = set(tuple(at_f[i]))
                c = g & b
                if c:
                    kol = len(c)
                    m = list(c)
                    for j in range (0, kol):
                        #st = at_f[4].replace(m[j], "")
                        #at_f[4] = st
                        #st = at_f[i].replace(m[j], "")
                        #at_f[i] = st
                        at_f[4].remove(m[j])
                        at_f[i].remove(m[j])
                        semester_number = m[j]
                        if i == 0:
                            attestation_form_id = 6
                        if i == 1:
                            attestation_form_id = 4
                        if i == 2:
                            attestation_form_id = 7
                        
                        cur.execute("SELECT syllabus_content_id FROM syllabus_content WHERE semester_number = %s AND attestation_form_id = %s AND discipline_id = %s AND syllabus_id = %s",
                            (semester_number, attestation_form_id, discipline_id, syllabus_id,))
                        a = cur.fetchall()
                        if not a:
                            cur.execute("SELECT COUNT(*) FROM syllabus_content")
                            kolvo = cur.fetchall()
                            syllabus_content_id = kolvo[0][0] + 1
                            cur.execute("INSERT INTO syllabus_content (syllabus_content_id, semester_number, attestation_form_id, discipline_id, syllabus_id  ) VALUES (%s, %s, %s, %s, %s)",
                                (syllabus_content_id, semester_number, attestation_form_id, discipline_id, syllabus_id ))
                            conn.commit()
           
                        
        for v in range (0, 3):
            if at_f[v]:
                if at_f[v][0]!='':
                    kol = len(at_f[v])
                    for j in range (0, kol):
                        semester_number = at_f[v][j]
                        attestation_form_id = v + 1
                        cur.execute("SELECT syllabus_content_id FROM syllabus_content WHERE semester_number = %s AND attestation_form_id = %s AND discipline_id = %s AND syllabus_id = %s",
                            (semester_number, attestation_form_id, discipline_id, syllabus_id,))
                        a = cur.fetchall()
                        if not a:
                            cur.execute("SELECT COUNT(*) FROM syllabus_content")
                            kolvo = cur.fetchall()
                            syllabus_content_id = kolvo[0][0] + 1
                            cur.execute("INSERT INTO syllabus_content (syllabus_content_id, semester_number, attestation_form_id, discipline_id, syllabus_id  ) VALUES (%s, %s, %s, %s, %s)",
                                (syllabus_content_id, semester_number, attestation_form_id, discipline_id, syllabus_id ))
                            conn.commit()

                
    

    ii+=1
    enter_in_db = True

#ДИСЦИПЛИНА И СОДЕРЖАНИЕ УЧЕБНОГО ПЛАНА
cur.close()
conn.close()


